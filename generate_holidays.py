import requests
from datetime import datetime, timedelta, date
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font
from collections import defaultdict

TZ = ZoneInfo("Europe/Paris")
LAT = 48.8534
LON = 2.4727
TODAY = date(2026, 5, 28)
END_YEAR = 2046
MARGIN_MINUTES = 6

HEBCAL_URL = "https://www.hebcal.com/hebcal"


def fetch_year(year):
    params = {
        "v": 1,
        "cfg": "json",
        "year": year,
        "month": "x",
        "maj": "on",
        "min": "off",
        "mod": "off",
        "nx": "off",
        "ss": "off",
        "mf": "off",
        "c": "on",
        "geo": "pos",
        "latitude": LAT,
        "longitude": LON,
        "tzid": "Europe/Paris",
        "M": "on",
        "s": "off",
        "i": "off",
        "lg": "fr",
    }
    resp = requests.get(HEBCAL_URL, params=params)
    resp.raise_for_status()
    return resp.json()


def build_blocks(all_data):
    all_items = []
    for data in all_data:
        all_items.extend(data.get("items", []))
    all_items.sort(key=lambda x: x.get("date", ""))

    yomtov_dates = {}
    candles_by_date = {}
    havdalah_by_date = {}

    for item in all_items:
        d = item["date"][:10]
        cat = item.get("category", "")
        if cat == "holiday" and item.get("yomtov", False):
            yomtov_dates[d] = item.get("title", "")
        elif cat == "candles":
            candles_by_date[d] = item["date"]
        elif cat == "havdalah":
            havdalah_by_date[d] = item["date"]

    sorted_dates = sorted(yomtov_dates.keys())
    if not sorted_dates:
        return []

    blocks = []
    block_start = sorted_dates[0]
    block_end = sorted_dates[0]
    block_names = [yomtov_dates[sorted_dates[0]]]

    for i in range(1, len(sorted_dates)):
        d = sorted_dates[i]
        prev = sorted_dates[i - 1]
        prev_date = datetime.strptime(prev, "%Y-%m-%d").date()
        cur_date = datetime.strptime(d, "%Y-%m-%d").date()
        gap = (cur_date - prev_date).days

        if gap <= 1:
            block_end = d
            block_names.append(yomtov_dates[d])
        elif gap == 2:
            between = (prev_date + timedelta(days=1)).strftime("%Y-%m-%d")
            between_weekday = (prev_date + timedelta(days=1)).weekday()
            if between_weekday == 5:
                block_end = d
                block_names.append("Chabbat")
                block_names.append(yomtov_dates[d])
            else:
                blocks.append((block_start, block_end, block_names))
                block_start = d
                block_end = d
                block_names = [yomtov_dates[d]]
        else:
            blocks.append((block_start, block_end, block_names))
            block_start = d
            block_end = d
            block_names = [yomtov_dates[d]]

    blocks.append((block_start, block_end, block_names))

    results = []
    for block_start, block_end, names in blocks:
        start_d = datetime.strptime(block_start, "%Y-%m-%d").date()
        end_d = datetime.strptime(block_end, "%Y-%m-%d").date()

        erev = (start_d - timedelta(days=1)).strftime("%Y-%m-%d")
        start_iso = candles_by_date.get(erev) or candles_by_date.get(block_start)

        day_after = (end_d + timedelta(days=1)).strftime("%Y-%m-%d")
        end_iso = havdalah_by_date.get(day_after) or havdalah_by_date.get(block_end)

        end_weekday = end_d.weekday()
        if end_weekday == 4:
            shabbat = (end_d + timedelta(days=1)).strftime("%Y-%m-%d")
            shabbat_end = (end_d + timedelta(days=2)).strftime("%Y-%m-%d")
            if shabbat_end in havdalah_by_date:
                end_iso = havdalah_by_date[shabbat_end]
                names = names + ["Chabbat"]
            elif shabbat in havdalah_by_date:
                end_iso = havdalah_by_date[shabbat]
                names = names + ["Chabbat"]

        start_weekday = start_d.weekday()
        if start_weekday == 0:
            prev_shabbat_candles = (start_d - timedelta(days=2)).strftime("%Y-%m-%d")
            if prev_shabbat_candles in candles_by_date:
                actual_start = candles_by_date[prev_shabbat_candles]
                if actual_start < (start_iso or "9999"):
                    start_iso = actual_start
                    names = ["Chabbat"] + names

        seen = set()
        unique = []
        for n in names:
            if n not in seen:
                seen.add(n)
                unique.append(n)

        results.append({
            "names": unique,
            "start_iso": start_iso,
            "end_iso": end_iso,
            "start_date": block_start,
            "end_date": block_end,
        })

    return results


def parse_iso_to_local(iso_str, margin_minutes=0):
    if not iso_str:
        return None
    return datetime.fromisoformat(iso_str).astimezone(TZ) + timedelta(minutes=margin_minutes)


def main():
    print("Récupération des données depuis Hebcal...")
    all_data = []
    for year in range(2026, END_YEAR + 1):
        print(f"  {year}...")
        all_data.append(fetch_year(year))

    print("Construction des blocs de fêtes...")
    blocks = build_blocks(all_data)

    filtered = [b for b in blocks
                if datetime.strptime(b["end_date"], "%Y-%m-%d").date() >= TODAY]

    print(f"{len(filtered)} blocs trouvés de {TODAY} à fin {END_YEAR}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Fêtes Juives"

    ws["A1"] = "Date et heure de début"
    ws["B1"] = "Date et heure de fin"
    ws["C1"] = "Nom de la fête"
    for col in ["A", "B"]:
        ws.column_dimensions[col].width = 22
    ws.column_dimensions["C"].width = 55
    for cell in [ws["A1"], ws["B1"], ws["C1"]]:
        cell.font = Font(bold=True)

    dt_fmt = "dd/mm/yyyy hh:mm"

    for i, b in enumerate(filtered, start=2):
        start_local = parse_iso_to_local(b["start_iso"], -MARGIN_MINUTES)
        end_local = parse_iso_to_local(b["end_iso"], +MARGIN_MINUTES)

        if start_local:
            ws.cell(row=i, column=1, value=start_local.replace(tzinfo=None))
            ws.cell(row=i, column=1).number_format = dt_fmt
        if end_local:
            ws.cell(row=i, column=2, value=end_local.replace(tzinfo=None))
            ws.cell(row=i, column=2).number_format = dt_fmt

        ws.cell(row=i, column=3, value=" + ".join(b["names"]))

    output = "fetes_juives_2026_2046.xlsx"
    wb.save(output)
    print(f"\nFichier généré : {output}")
    print(f"{len(filtered)} lignes")

    print("\nAperçu :")
    for b in filtered:
        s = parse_iso_to_local(b["start_iso"], -MARGIN_MINUTES)
        e = parse_iso_to_local(b["end_iso"], +MARGIN_MINUTES)
        s_str = s.strftime("%d/%m/%Y %H:%M") if s else "?"
        e_str = e.strftime("%d/%m/%Y %H:%M") if e else "?"
        print(f"  {s_str} → {e_str}  |  {' + '.join(b['names'])}")


if __name__ == "__main__":
    main()
